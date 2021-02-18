"""
Arquivo com as funções que fazem a leitura do arquivo do Excel,
manipulam ele e o deixam pronto para ser tratado.
"""
from openpyxl import load_workbook
from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from typing import List

# *
# Funções que cuidam de carregar o arquivo do Excel
# *

def load_workbook_file(filename: str) -> Workbook:
    """
    Função que carrega o arquivo na memória.

    Arguments:
        filename (str): caminho absoluto para o arquivo .xlsx ou apenas nome, 
                        caso esteja localizado na raiz do projeto

    Return:
        Workbook: arquivo para ser manipulado pelo módulo 
    """
    print('Carregando arquivo...')
    workbook_file = load_workbook(filename=filename)
    
    print('Arquivo carregado com sucesso.')
    return workbook_file


# *
# Funções que fazem verificações e retornam valores booleanos
# *

def verify_empty_row(row: tuple) -> bool:
    """
    Função que verifica se a linha está vazia.

    Arguments:
        row (tuple): tupla com os valores armazenados na linha

    Arguments:
        bool: indica se a linha está ou não vazia
    """

    # A LINHA VAI SER CONSIDERADA VAZIA QUANDO AS 
    # COLUNAS A, B, F ESTIVEREM VAZIAS AO MESMO TEMPO
    is_empty_row = (row[0] == None) and (row[1] == None) and (row[6] == None)

    return is_empty_row


def verify_row_contains_rating(sheet_row: tuple) -> bool:
    """
    Função que retorna apenas as linhas onde o campo de `avaliação
    concreta` é marcado com `SIM`.

    Arguments:
        sheet_row (tuple): tupla de valores guardados na linha
    
    Return:
        bool: indica se a linha contém ou não avaliação
    """
    rating_cell = sheet_row[6]
    contains_rating = False

    if rating_cell and (rating_cell.lower() == 'x'):
        contains_rating = True

    return contains_rating


# *
# Funções que executam as etapas iniciais da preparação do arquivo 
# de planilha como:
#       - contar as linhas utilizadas
#       - obter as linhas com o campo `avaliação concreta` marcado com `SIM`
#       - formatar as linhas para o formato que será utilizado no final
# *

def get_worksheet_used_rows(sheet: Worksheet) -> int:
    """
    Função que conta as linhas utilizadas da planilha.

    Arguments:
        sheet (Worksheet): objeto da planilha
    
    Return:
        int: quantidade de linhas utilizadas
    """
    print(f'Contando linhas utilizadas na planilha "{sheet.title}"...')
    count = 0

    for row in sheet.iter_rows(values_only=True):
        if verify_empty_row(row):
            break

        count += 1

    return count 


def get_rating_rows(sheet: Worksheet, max_row: int) -> List[tuple]:
    """
    Função que retorna todas as linhas onde o campo de `avaliação
    concreta` é marcado com `SIM`.

    Arguments:
        sheet (Worksheet): planilha que será analisada
        max_row (int): máximo de linhas que serão processadas
    
    Return:
        List[tuple]: lista com todas as linhas que contém avaliação
    """
    print('Verificando as linhas da planilha...')
    rated_rows = []

    for index, row in enumerate(sheet.iter_rows(values_only=True)):
        
        if (index < max_row) and verify_row_contains_rating(row):
            rated_rows.append(row)

    return rated_rows


def get_policy_level(rated_row: tuple) -> str:
    """
    Função que retorna o valor do campo de `política avaliada`.

    Arguments:
        rated_row (tuple): linha que possui `avaliação concreta`
    
    Return:
        str: valor da política
    """
    policy = None

    municipal_col = rated_row[10] 
    estadual_col = rated_row[11]
    federal_col = rated_row[12]

    if municipal_col:
        if municipal_col.lower() == 'internacional':
            policy = 'Internacional'

        elif municipal_col.lower() == 'x':
            policy = 'Municipal'

    elif estadual_col:
        if estadual_col.lower() == 'internacional':
            policy = 'Internacional'

        elif estadual_col.lower() == 'x':
            policy = 'Estadual'

    elif federal_col:
        if federal_col.lower() == 'internacional':
            policy = 'Internacional'

        elif federal_col.lower() == 'x':
            policy = 'Federal'

    return policy


def get_formated_items(rated_rows: List[tuple], sheet_title: str) -> List[tuple]:
    """
    Função que formata as linhas que possuem `avaliação concreta`
    para o formato final adotado.

    Arguments:
        rated_rows (List[tuple]): lista de linhas que possuem `avaliação concreta`
        sheet_title (str): titulo da planilha, encontrado na célula `A1`
    
    Return:
        List[tuple]: lista com as linhas utilizando o formato final
    """
    print('Formatando a planilha...')
    formated_items = []

    for row in rated_rows:
        policy = get_policy_level(row)

        formated_item = (
            sheet_title,
            row[1],
            row[2],
            row[3],
            policy
            )
        formated_items.append(formated_item)
    
    return formated_items


# *
# Função que executa os passos em ordem lógica
# *

def read_and_format_workbook(filename: str) -> List[list]:
    """
    Função que executa os passos de leitura e formatação em ordem lógica.
    Os passos são:
        - carregar arquivo
        - obter nomes das planílhas
        - iterando sobre cada planilha:
            - obter o título do grupo
            - obter o número de linhas utilizadas 
            - verificar as linhas com avaliação
            - formatar as linhas para a saída padrão

    Arguments:
        filename (str): caminho absoluto para a planilha ou apenas nome,
                        caso ela esteja na raiz do projeto

    Returns:
        List[list]: lista com todas as linhas que contém avaliação 
                    já formatadas
    """
    workbook = load_workbook_file(filename)
    sheets_names = workbook.sheetnames

    formatted_sheets = []

    for sheet_name in sheets_names:
        sheet = workbook[sheet_name]
        sheet_title = sheet['A1'].value

        sheet_rows = get_worksheet_used_rows(sheet)
        sheet_rated = get_rating_rows(sheet, sheet_rows)
        sheet_formated = get_formated_items(sheet_rated, sheet_title)

        formatted_sheets += sheet_formated
        print()

    return formatted_sheets
