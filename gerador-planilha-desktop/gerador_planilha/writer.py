from openpyxl.workbook import Workbook
from openpyxl.worksheet.worksheet import Worksheet

from typing import List

from . import styles
from . import window_updater


# *
# Funções que povoam/escrevem partes do arquivo de planilha
# *

def write_column_titles(workbook: Workbook, window=None) -> Workbook:
    """
    Função que escreve os títulos das colunas no objeto
    de livro de planilhas.

    Arguments:
        workbook (Workbook): objeto de livro de planilhas

    Return:
        Workbook: livro de planilhas com os títulos das colunas
    """
    sheet = workbook.active
    col_titles = styles.title_columns
    
    # DEFININDO LARGURA DAS COLUNAS
    sheet.column_dimensions['A'].width = 8
    sheet.column_dimensions['B'].width = 25
    sheet.column_dimensions['C'].width = 30
    sheet.column_dimensions['D'].width = 8
    sheet.column_dimensions['E'].width = 20
    sheet.column_dimensions['F'].width = 25
    sheet.column_dimensions['G'].width = 10
    sheet.column_dimensions['H'].width = 30
    sheet.column_dimensions['I'].width = 30
    sheet.column_dimensions['J'].width = 30
    sheet.column_dimensions['K'].width = 30

    print('Gravando os títulos das colunas....')
    window_updater.update(window)

    for col, col_title in enumerate(col_titles):
        sheet.cell(1, col + 1).value = col_title

        if col_title is not None:
            sheet.cell(1, col + 1).font = styles.title_column_font
            sheet.cell(1, col + 1).fill = styles.title_column_fill
            sheet.cell(1, col + 1).border = styles.title_column_border
            sheet.cell(1, col + 1).alignment = styles.title_column_alignment

    return workbook


def write_formated_data(workbook: Workbook, row_list: List[tuple], window=None) -> Workbook:
    """
    Função que escreve os dados já no formato padrão no objeto
    de livro de planilhas.

    Arguments:
        workbook (Workbook): objeto de livro de planilhas
        row_list (List[tuple]): lista com as linhas formatadas

    Return:
        Workbook: livro de planilhas com os dados gravados
    """
    sheet = workbook.active

    print('Inserindo os dados no arquivo...')
    window_updater.update(window)

    for row, row_content in enumerate(row_list):

        # INSERE O NÚMERO NA PRIMEIRA COLUNA DA LINHA
        sheet.cell(row + 2, 1).value = row + 1
        sheet.cell(row + 2, 1).border = styles.content_border
        sheet.cell(row + 2, 1).alignment = styles.content_align_center

        # INSERE OS DADOS NAS COLUNAS RESTANTES
        for col, col_content in enumerate(row_content):
            sheet.cell(row + 2, col + 2).value = col_content

            if col_content is not None:
                sheet.cell(row + 2, col + 2).border = styles.content_border

                # QUANDO A COLUNA NÃO FOR DE OBRA, O TEXTO DEVE ESTAR CENTRALIZADO
                if col != 1:
                    sheet.cell(row + 2, col + 2).alignment = styles.content_align_center

                else:
                    # QUANDO A COLUNA FOR DE OBRA, O TEXTO NÃO ESTARÁ CENTRALIZADO
                    sheet.cell(row + 2, col + 2).alignment = styles.content_align_no_center

    return workbook


# *
# Funções que aplicam preparam e geram o arquivo de planilhas
# *

def create_new_workbook_from_row_list(row_list: List[tuple], window=None) -> Workbook:
    """
    Função que cria o objeto de livro de planilhas e executa as funções
    de povoamento e estilização em ordem lógica antes de exportar para um 
    arquivo.

    Arguments:
        row_list (List[tuple]): lista com as linhas formatadas

    Return:
        Workbook: livro de planilhas com os dados gravados e estilizados
    """
    workbook = Workbook()

    write_column_titles(workbook, window=window)
    write_formated_data(workbook, row_list, window=window)

    return workbook


def save_workbook_file(workbook: Workbook, filename: str, window=None):
    """
    Função que exporta um livro de planilhas para um arquivo
    no sistema de arquivos.

    Arguments:
        workbook (Workbook): livro de planilhas com os dados gravados e estilizados
        filename (str): caminho absoluto para o arquivo ou apenas nome,
                        caso o arquivo esteja na raiz do projeto
    """
    print('Salvando arquivo no sistema...')
    window_updater.update(window)

    workbook.save(filename=filename)
    
    print('Arquivo salvo com sucesso!')
    window_updater.update(window)


def open_workbook_file(filename: str, window=None):
    """
    Função que abre o arquivo passado com o programa padrão definido no
    sistema operacional.

    Arguments:
        filename (str): caminho absoluto para o arquivo ou apenas nome,
                        caso o arquivo esteja na raiz do projeto
    """
    from platform import system as system_name
    from os import system as execute

    print('Abrindo arquivo...')
    window_updater.update(window)

    if system_name() == 'Linux':
        execute(f'xdg-open "{filename}"')

    elif system_name() == 'Windows':
        execute(f'start "{filename}"')

    elif system_name() == 'Darwin':
        execute(f'open "{filename}"')



# *
# Funções que executam os passos de escrita em ordem lógica
# *

def generate_workbook_file(row_list: List[tuple], filename: str, window=None):
    """
    Função que executa todos os passos para gerar o arquivo
    de planilhas no sistema. Os passos são:
        - criar um objeto de planilha
        - gravar os dados no objeto
        - salvar o arquivo da planilha no sistema

    Arguments:
        row_list (List[tuple]): lista com as linhas formatadas
        filename (str): caminho absoluto para o arquivo ou apenas nome,
                        caso o arquivo esteja na raiz do projeto
    """
    workbook = create_new_workbook_from_row_list(row_list, window=window)
    save_workbook_file(workbook, filename, window=window)
    open_workbook_file(filename, window=window)
